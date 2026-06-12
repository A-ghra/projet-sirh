"""Importation de contrats — CSV, Excel, association employé par matricule."""
import csv
import io
import re
from datetime import datetime

import openpyxl

from .models import Contract, Employee
from .contract_service import generate_contract_number


FIELD_ALIASES = {
    'matricule': 'matricule', 'employee': 'matricule', 'employe': 'matricule',
    'type': 'contract_type', 'contract_type': 'contract_type', 'typecontrat': 'contract_type',
    'debut': 'start_date', 'start': 'start_date', 'start_date': 'start_date', 'datedebut': 'start_date',
    'fin': 'end_date', 'end': 'end_date', 'end_date': 'end_date', 'datefin': 'end_date',
    'salaire': 'salary_base', 'salary': 'salary_base', 'salary_base': 'salary_base',
    'poste': 'position_title', 'position': 'position_title',
    'numero': 'contract_number', 'contract_number': 'contract_number', 'numero_contrat': 'contract_number',
}


def _normalize_header(h):
    return re.sub(r'[^a-z0-9]', '', (h or '').lower())


def _parse_date(val):
    if not val:
        return None
    if hasattr(val, 'date'):
        return val.date() if hasattr(val, 'date') else val
    s = str(val).strip()[:10]
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _map_row(row_dict):
    data = {}
    for key, val in row_dict.items():
        norm = _normalize_header(key)
        for alias, field in FIELD_ALIASES.items():
            if _normalize_header(alias) == norm:
                data[field] = val
                break
    if data.get('start_date'):
        data['start_date'] = _parse_date(data['start_date'])
    if data.get('end_date'):
        data['end_date'] = _parse_date(data['end_date'])
    if data.get('salary_base'):
        try:
            data['salary_base'] = float(str(data['salary_base']).replace(',', '.').replace(' ', ''))
        except ValueError:
            data['salary_base'] = 0
    return data


def import_contracts_from_rows(rows, user=None, file_obj=None):
    created, errors = [], []
    for i, row in enumerate(rows, start=2):
        data = _map_row(row)
        mat = data.get('matricule')
        if not mat:
            errors.append({'row': i, 'error': 'Matricule manquant'})
            continue
        emp = Employee.objects.filter(matricule__iexact=str(mat).strip()).first()
        if not emp:
            errors.append({'row': i, 'error': f'Employé introuvable : {mat}'})
            continue
        if not data.get('start_date'):
            errors.append({'row': i, 'error': 'Date de début manquante'})
            continue
        try:
            c = Contract.objects.create(
                employee=emp,
                contract_number=data.get('contract_number') or generate_contract_number(emp),
                contract_type=data.get('contract_type') or emp.contract_type or 'CDI',
                start_date=data['start_date'],
                end_date=data.get('end_date'),
                salary_base=data.get('salary_base') or emp.salary_base,
                position_title=data.get('position_title') or emp.position,
                status='DRAFT',
                created_by=user,
            )
            if file_obj and i == 2:
                c.file = file_obj
                c.save(update_fields=['file'])
            created.append(c.id)
        except Exception as e:
            errors.append({'row': i, 'error': str(e)})
    return {'created': len(created), 'ids': created, 'errors': errors}


def import_from_csv(file_obj, user=None):
    text = file_obj.read()
    if isinstance(text, bytes):
        text = text.decode('utf-8-sig', errors='replace')
    reader = csv.DictReader(io.StringIO(text))
    return import_contracts_from_rows(reader, user=user)


def import_from_excel(file_obj, user=None):
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return import_contracts_from_rows(rows, user=user, file_obj=file_obj)
